from pandas import DataFrame, Series
from openpyxl import load_workbook
import glob, sys
import pandas as pd

path				= ''
filelist = glob.glob("*.xls")

## tables				= {}

## for filename in filelist:
##	xls_file		= pd.ExcelFile(filename)
##	sheetnames		= xls_file.sheet_names
##	for sheet in sheetnames:
##		data		= xls_file.parse(sheet)
##		dataname	= str(filename) + "-" + str(sheet)
##		tables[dataname]= data

filename			= path + 'RAW Charter Report Card.xls'
xls_file			= pd.ExcelFile(filename)
charters			= xls_file.parse('COMMSCHL', index_col=0)

filename			= path + 'RAW Charter Teacher Data.xls'
xls_file			= pd.ExcelFile(filename)
charter_teacher			= xls_file.parse('TEACHER', index_col=0)


filename			= path + 'RAW District to Charter Transfer by Performance Data.xlsx'
funding_xlsx			= load_workbook(filename)
funding_sheet			= funding_xlsx['Sheet1']
cell_range			= funding_sheet.range('A1:CA10000')

public_funding			= {}
irn_col				= False

for row in cell_range:
	if not(irn_col):
		print '1'
		try:
			irn_col		= row.index('IRN')
			transfer_col	= row.index('TRANSFER')
		except:
			pass
	else:
		print '2'
		if public_funding[row(irn_col)] in public_funding:
			public_funding[row(irn_col)] = \
				public_funding[row(irn_col)] + row(transfer_col)	
		else:
			public_funding[row(irn_col)] = row(transfer_col)

#print public_funding
sys.exit()

#charters			= tables['RAW Charter Report Card.xls-COMMSCHL']

cols				= charters.columns.values
cols.remove('Building IRN')
cols.remove('Building Name')
cols.remove('District IRN')
cols.remove('District Name')
cols.remove('County')
cols.remove('Region')
cols.remove('Street address')
cols.remove('City and Zip code')
cols.remove('Letter grade of performance index')
charters.drop(cols,inplace=True,axis=1)
charters.rename(columns={'Building IRN': 'School IRN', 'Building Name': 'School Name'}, inplace=True)

charters['Average Years of teacher experience']	= charter_teacher['Average Years of teacher experience']

## Classroom Budget
## Public Funding 


